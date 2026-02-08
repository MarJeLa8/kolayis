"""
Import servisi - CSV ve Excel dosyalarindan veri aktarimi.
Musteri, urun ve fatura import islemlerini yonetir.
"""
import csv
import uuid
from io import StringIO, BytesIO
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from sqlalchemy.orm import Session

from kolayis.models.customer import Customer
from kolayis.models.invoice import Invoice, InvoiceItem
from kolayis.schemas.customer import CustomerCreate
from kolayis.schemas.product import ProductCreate
from kolayis.schemas.invoice import InvoiceCreate, InvoiceItemCreate
from kolayis.services import customer as customer_service
from kolayis.services import product as product_service
from kolayis.services import invoice as invoice_service
from kolayis.services.activity import log_activity


def parse_file(filename: str, raw_bytes: bytes) -> list[list[str]]:
    """
    CSV veya Excel dosyasini parse et.
    Dosya uzantisina gore otomatik algilar.
    Dondurur: satir listesi (her satir = hucre listesi)
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "xlsx":
        return _parse_excel(raw_bytes)
    else:
        return _parse_csv(raw_bytes)


def _parse_csv(raw_bytes: bytes) -> list[list[str]]:
    """CSV dosyasini parse et (UTF-8 BOM + latin-1 fallback)."""
    # UTF-8 BOM varsa kaldir
    if raw_bytes.startswith(b'\xef\xbb\xbf'):
        raw_bytes = raw_bytes[3:]

    try:
        content = raw_bytes.decode("utf-8")
    except UnicodeDecodeError:
        content = raw_bytes.decode("latin-1")

    reader = csv.reader(StringIO(content))
    return list(reader)


def _parse_excel(raw_bytes: bytes) -> list[list[str]]:
    """Excel (xlsx) dosyasini parse et."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        # None degerleri bos stringe cevir
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()
    return rows


def _parse_date(value: str) -> date:
    """Tarih stringini parse et. Desteklenen formatlar: YYYY-MM-DD, DD.MM.YYYY, DD/MM/YYYY"""
    value = value.strip()
    if not value:
        raise ValueError("Tarih bos olamaz")

    # YYYY-MM-DD
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Gecersiz tarih formati: {value} (Beklenen: YYYY-MM-DD veya DD.MM.YYYY)")


def import_customers(db: Session, owner_id: uuid.UUID, rows: list[list[str]]) -> dict:
    """
    Musteri satirlarini iceri aktar.
    rows[0] = baslik satiri (atlanir), rows[1:] = veri satirlari
    Kolon sirasi: Sirket Adi, Ilgili Kisi, Email, Telefon, Adres, Vergi No
    """
    if len(rows) < 2:
        return {"total": 0, "success": 0, "failed": 0, "errors": []}

    data_rows = rows[1:]
    success_count = 0
    failed_count = 0
    errors = []

    for idx, row in enumerate(data_rows, start=2):
        # Bos satirlari atla
        if not row or all(cell.strip() == "" for cell in row):
            continue

        try:
            company_name = row[0].strip() if len(row) > 0 else ""
            contact_name = row[1].strip() if len(row) > 1 else ""
            email = row[2].strip() if len(row) > 2 else ""
            phone = row[3].strip() if len(row) > 3 else ""
            address = row[4].strip() if len(row) > 4 else ""
            tax_number = row[5].strip() if len(row) > 5 else ""

            if not company_name:
                raise ValueError("Sirket adi bos olamaz")

            data = CustomerCreate(
                company_name=company_name,
                contact_name=contact_name or None,
                email=email or None,
                phone=phone or None,
                address=address or None,
                tax_number=tax_number or None,
            )
            customer_service.create_customer(db, owner_id, data)
            success_count += 1
        except Exception as e:
            failed_count += 1
            errors.append({"row": idx, "message": str(e)})

    return {
        "total": success_count + failed_count,
        "success": success_count,
        "failed": failed_count,
        "errors": errors,
    }


def import_products(db: Session, owner_id: uuid.UUID, rows: list[list[str]]) -> dict:
    """
    Urun satirlarini iceri aktar.
    Kolon sirasi: Urun Adi, Aciklama, Birim Fiyat, Birim, KDV Orani, Stok
    """
    if len(rows) < 2:
        return {"total": 0, "success": 0, "failed": 0, "errors": []}

    data_rows = rows[1:]
    success_count = 0
    failed_count = 0
    errors = []

    for idx, row in enumerate(data_rows, start=2):
        if not row or all(cell.strip() == "" for cell in row):
            continue

        try:
            name = row[0].strip() if len(row) > 0 else ""
            description = row[1].strip() if len(row) > 1 else ""
            unit_price_str = row[2].strip() if len(row) > 2 else ""
            unit = row[3].strip() if len(row) > 3 else "adet"
            tax_rate_str = row[4].strip() if len(row) > 4 else "20"
            stock_str = row[5].strip() if len(row) > 5 else ""

            if not name:
                raise ValueError("Urun adi bos olamaz")
            if not unit_price_str:
                raise ValueError("Birim fiyat bos olamaz")

            try:
                unit_price = Decimal(unit_price_str.replace(",", "."))
            except InvalidOperation:
                raise ValueError(f"Gecersiz birim fiyat: {unit_price_str}")

            try:
                tax_rate = int(tax_rate_str) if tax_rate_str else 20
            except ValueError:
                raise ValueError(f"Gecersiz KDV orani: {tax_rate_str}")

            stock = None
            if stock_str:
                try:
                    stock = int(stock_str)
                except ValueError:
                    raise ValueError(f"Gecersiz stok degeri: {stock_str}")

            data = ProductCreate(
                name=name,
                description=description or None,
                unit_price=unit_price,
                unit=unit or "adet",
                tax_rate=tax_rate,
                stock=stock,
            )
            product_service.create_product(db, owner_id, data)
            success_count += 1
        except Exception as e:
            failed_count += 1
            errors.append({"row": idx, "message": str(e)})

    return {
        "total": success_count + failed_count,
        "success": success_count,
        "failed": failed_count,
        "errors": errors,
    }


def import_invoices(db: Session, owner_id: uuid.UUID, rows: list[list[str]]) -> dict:
    """
    Fatura satirlarini iceri aktar.
    Her satir bir fatura kalemi. Ayni musteri_adi + fatura_tarihi = ayni fatura.
    Kolon sirasi: Musteri Adi, Fatura Tarihi, Vade Tarihi, Kalem Aciklama, Miktar, Birim Fiyat, KDV Orani
    """
    if len(rows) < 2:
        return {"total": 0, "success": 0, "failed": 0, "errors": []}

    data_rows = rows[1:]
    errors = []

    # 1. Asamada: satirlari grupla (musteri_adi + fatura_tarihi = bir fatura)
    # Ayrica satir bazli dogrulama yap
    invoice_groups = {}  # key: (musteri_adi, fatura_tarihi_str) -> items listesi

    for idx, row in enumerate(data_rows, start=2):
        if not row or all(cell.strip() == "" for cell in row):
            continue

        try:
            musteri_adi = row[0].strip() if len(row) > 0 else ""
            fatura_tarihi_str = row[1].strip() if len(row) > 1 else ""
            vade_tarihi_str = row[2].strip() if len(row) > 2 else ""
            kalem_aciklama = row[3].strip() if len(row) > 3 else ""
            miktar_str = row[4].strip() if len(row) > 4 else ""
            birim_fiyat_str = row[5].strip() if len(row) > 5 else ""
            kdv_orani_str = row[6].strip() if len(row) > 6 else "20"

            # Zorunlu alan kontrolleri
            if not musteri_adi:
                raise ValueError("Musteri adi bos olamaz")
            if not fatura_tarihi_str:
                raise ValueError("Fatura tarihi bos olamaz")
            if not kalem_aciklama:
                raise ValueError("Kalem aciklama bos olamaz")
            if not miktar_str:
                raise ValueError("Miktar bos olamaz")
            if not birim_fiyat_str:
                raise ValueError("Birim fiyat bos olamaz")

            # Tarih dogrulama
            fatura_tarihi = _parse_date(fatura_tarihi_str)
            vade_tarihi = _parse_date(vade_tarihi_str) if vade_tarihi_str else None

            # Sayisal deger dogrulama
            try:
                miktar = Decimal(miktar_str.replace(",", "."))
            except InvalidOperation:
                raise ValueError(f"Gecersiz miktar: {miktar_str}")

            try:
                birim_fiyat = Decimal(birim_fiyat_str.replace(",", "."))
            except InvalidOperation:
                raise ValueError(f"Gecersiz birim fiyat: {birim_fiyat_str}")

            try:
                kdv_orani = int(kdv_orani_str) if kdv_orani_str else 20
            except ValueError:
                raise ValueError(f"Gecersiz KDV orani: {kdv_orani_str}")

            # Gruplama anahtari
            key = (musteri_adi, fatura_tarihi_str)
            if key not in invoice_groups:
                invoice_groups[key] = {
                    "musteri_adi": musteri_adi,
                    "fatura_tarihi": fatura_tarihi,
                    "vade_tarihi": vade_tarihi,
                    "items": [],
                    "rows": [],  # hata raporlamasi icin satir numaralari
                }
            invoice_groups[key]["items"].append({
                "description": kalem_aciklama,
                "quantity": miktar,
                "unit_price": birim_fiyat,
                "tax_rate": kdv_orani,
            })
            invoice_groups[key]["rows"].append(idx)

        except Exception as e:
            errors.append({"row": idx, "message": str(e)})

    # 2. Asama: gruplari faturaya cevir
    success_count = 0
    failed_count = 0

    for key, group in invoice_groups.items():
        try:
            # Musteriyi bul (company_name ile)
            customer = db.query(Customer).filter(
                Customer.owner_id == owner_id,
                Customer.company_name.ilike(group["musteri_adi"]),
            ).first()

            if not customer:
                raise ValueError(f"Musteri bulunamadi: '{group['musteri_adi']}'")

            # Fatura kalemleri olustur
            items = []
            for item_data in group["items"]:
                items.append(InvoiceItemCreate(
                    description=item_data["description"],
                    quantity=item_data["quantity"],
                    unit_price=item_data["unit_price"],
                    tax_rate=item_data["tax_rate"],
                ))

            # Fatura olustur
            invoice_data = InvoiceCreate(
                customer_id=customer.id,
                invoice_date=group["fatura_tarihi"],
                due_date=group["vade_tarihi"],
                status="draft",
                items=items,
            )
            invoice_service.create_invoice(db, owner_id, invoice_data)
            success_count += 1

        except Exception as e:
            failed_count += 1
            # Gruptaki tum satirlara hata ekle
            row_nums = ", ".join(str(r) for r in group["rows"])
            errors.append({"row": row_nums, "message": str(e)})

    return {
        "total": success_count + failed_count,
        "success": success_count,
        "failed": failed_count,
        "errors": errors,
    }
